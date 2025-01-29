import openpyxl
import numpy as np

# 创建一个新的工作簿和工作表
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# 写入一些数据
data = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]

for row in data:
    ws.append(row)

# 保存工作簿
file_path = 'example.xlsx'
wb.save(file_path)

# 重新打开工作簿并读取数据
wb_read = openpyxl.load_workbook(file_path)
ws_read = wb_read['Sheet1']

# 读取数据并打印
read_data = []
for row in ws_read.iter_rows(values_only=True):
    read_data.append(row)

print(read_data)

# 检查是否有NaN值
for r in read_data:
    for value in r:
        if isinstance(value, float) and np.isnan(value):
            print("Found NaN value!")

# 清理资源
wb_read.close()